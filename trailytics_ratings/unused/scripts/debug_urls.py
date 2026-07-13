"""Investigate URL patterns and debug brand detection."""
import sys, json
sys.path.insert(0, 'ml_pipeline')
import openpyxl

wb = openpyxl.load_workbook('dist/assets/prestige_comp_reviews_data.xlsx', read_only=True)

# Check 20 sample URLs from each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if count >= 15:
            break
        url = str(row[5]) if row[5] and str(row[5]) != '\\N' else ''
        pid = str(row[4]) if row[4] else ''
        title = str(row[7]) if row[7] and str(row[7]) != '\\N' else ''
        
        if url and len(url) > 10:
            print(f"\n  PID: {pid}")
            print(f"  URL: {url[:120]}")
            print(f"  Title: {title[:80]}")
            count += 1

wb.close()
