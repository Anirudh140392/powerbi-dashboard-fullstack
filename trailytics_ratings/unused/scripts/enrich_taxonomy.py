"""
Phase 2: Keyword Enrichment from 70K Real Competitor Reviews
=============================================================
Reads the Excel file, extracts all review text, runs n-gram analysis,
and enriches the ML taxonomy with newly discovered keywords.
"""

import os
import sys
import re
import json
import math
from collections import Counter, defaultdict
from pathlib import Path

# Add project root
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "ml_pipeline"))

import openpyxl

# ============================================================
# CONFIG
# ============================================================
EXCEL_PATH = PROJECT_ROOT / "dist" / "assets" / "prestige_comp_reviews_data.xlsx"
TAXONOMY_OUTPUT = PROJECT_ROOT / "ml_pipeline" / "enriched_keywords.json"
REPORT_OUTPUT = PROJECT_ROOT / "scripts" / "enrichment_report.json"

# Stop words for n-gram extraction (common English + Hindi transliterations)
STOP_WORDS = {
    "the", "a", "an", "is", "are", "was", "were", "be", "been", "being",
    "have", "has", "had", "do", "does", "did", "will", "would", "shall",
    "should", "may", "might", "must", "can", "could", "and", "but", "or",
    "nor", "not", "so", "yet", "for", "of", "at", "by", "in", "on", "to",
    "up", "it", "its", "this", "that", "these", "those", "my", "me", "we",
    "our", "you", "your", "he", "she", "they", "them", "his", "her",
    "i", "am", "with", "from", "as", "if", "then", "than", "too", "very",
    "just", "more", "also", "about", "all", "any", "each", "every", "both",
    "few", "other", "some", "such", "no", "only", "own", "same", "\\n",
    "read", "read more", "readmore", "\\\\n", "none", "na", "nil", "n/a",
    "product", "review", "reviews", "amazon", "flipkart", "bought", "buy",
    "purchase", "purchased", "got", "get", "getting", "one", "two", "three",
    "months", "month", "year", "years", "day", "days", "week", "weeks",
    "after", "before", "during", "since", "ago", "now", "still", "even",
    "much", "many", "lot", "lots", "bit", "little", "back", "again",
    "like", "well", "really", "make", "made", "use", "used", "using",
    "time", "first", "last", "new", "old", "long", "come", "came",
    "thing", "things", "way", "went", "going", "go", "take", "took",
    "see", "saw", "look", "looked", "give", "gave", "given",
}


def clean_text(text: str) -> str:
    """Clean and normalize review text."""
    if not text:
        return ""
    text = str(text)
    # Remove \\N placeholders
    text = text.replace("\\N", "").replace("\\n", " ")
    # Remove URLs
    text = re.sub(r'https?://\S+', '', text)
    # Remove HTML tags
    text = re.sub(r'<[^>]+>', '', text)
    # Remove special characters but keep alphanumeric, spaces, and basic punctuation
    text = re.sub(r'[^a-zA-Z0-9\s.,!?-]', ' ', text)
    # Normalize whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    return text.lower()


def extract_ngrams(text: str, n: int) -> list:
    """Extract n-grams from text, filtering stop words."""
    words = text.split()
    words = [w for w in words if len(w) > 1 and w not in STOP_WORDS]
    if len(words) < n:
        return []
    return [" ".join(words[i:i+n]) for i in range(len(words) - n + 1)]


def read_excel_reviews() -> list:
    """Read all reviews from both Excel sheets."""
    print(f"📂 Reading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True)
    
    all_reviews = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        count = 0
        platform = sheet_name.lower()
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            title = str(row[7]) if row[7] and str(row[7]) != "\\N" else ""
            body = str(row[8]) if row[8] and str(row[8]) != "\\N" else ""
            text = f"{title}. {body}".strip(". ")
            
            if not text or len(text) < 5:
                continue
            
            rating = 0
            try:
                rating = int(row[6]) if row[6] and str(row[6]) != "\\N" else 0
            except (ValueError, TypeError):
                pass
            
            all_reviews.append({
                "text": text,
                "rating": rating,
                "platform": platform,
                "pid": str(row[4]) if row[4] else "",
                "url": str(row[5])[:200] if row[5] else "",
            })
            count += 1
        
        print(f"   {sheet_name}: {count:,} reviews loaded")
    
    wb.close()
    print(f"   Total: {len(all_reviews):,} reviews")
    return all_reviews


def analyze_ngrams(reviews: list) -> dict:
    """Run comprehensive n-gram analysis on the review corpus."""
    print(f"\n🔬 Running n-gram analysis on {len(reviews):,} reviews...")
    
    # Split reviews by rating bucket
    buckets = {1: [], 2: [], 3: [], 4: [], 5: []}
    for r in reviews:
        rating = r["rating"]
        if rating in buckets:
            buckets[rating].append(clean_text(r["text"]))
    
    for rating, texts in buckets.items():
        print(f"   {rating}★: {len(texts):,} reviews")
    
    results = {}
    
    for n in [1, 2, 3]:
        label = {1: "unigrams", 2: "bigrams", 3: "trigrams"}[n]
        print(f"\n   Extracting {label}...")
        
        # Per-bucket n-gram counts
        bucket_ngrams = {}
        global_ngrams = Counter()
        
        for rating, texts in buckets.items():
            counter = Counter()
            for text in texts:
                ngrams = extract_ngrams(text, n)
                counter.update(ngrams)
            bucket_ngrams[rating] = counter
            global_ngrams.update(counter)
        
        # Find discriminative n-grams (high in 1-2★ OR 4-5★, low in opposite)
        negative_ngrams = Counter()
        positive_ngrams = Counter()
        
        for ngram, count in global_ngrams.items():
            if count < 5:  # Minimum frequency threshold
                continue
            
            neg_count = bucket_ngrams.get(1, Counter()).get(ngram, 0) + bucket_ngrams.get(2, Counter()).get(ngram, 0)
            pos_count = bucket_ngrams.get(4, Counter()).get(ngram, 0) + bucket_ngrams.get(5, Counter()).get(ngram, 0)
            neu_count = bucket_ngrams.get(3, Counter()).get(ngram, 0)
            
            # Normalize by bucket size
            neg_total = len(buckets.get(1, [])) + len(buckets.get(2, []))
            pos_total = len(buckets.get(4, [])) + len(buckets.get(5, []))
            
            if neg_total > 0 and pos_total > 0:
                neg_rate = neg_count / neg_total
                pos_rate = pos_count / pos_total
                
                if neg_rate > pos_rate * 1.5 and neg_count >= 10:
                    negative_ngrams[ngram] = neg_count
                elif pos_rate > neg_rate * 1.5 and pos_count >= 10:
                    positive_ngrams[ngram] = pos_count
        
        results[label] = {
            "total_unique": len(global_ngrams),
            "top_global": global_ngrams.most_common(100),
            "top_negative": negative_ngrams.most_common(100),
            "top_positive": positive_ngrams.most_common(100),
        }
        
        print(f"      Total unique: {len(global_ngrams):,}")
        print(f"      Negative-skewed: {len(negative_ngrams)}")
        print(f"      Positive-skewed: {len(positive_ngrams)}")
    
    return results


def map_ngrams_to_taxonomy(ngram_results: dict) -> dict:
    """Map discovered n-grams into taxonomy categories."""
    print("\n📋 Mapping n-grams to taxonomy categories...")
    
    # Category keyword patterns (for auto-mapping discovered n-grams)
    CATEGORY_PATTERNS = {
        "Quality": {
            "Lid_Issues": ["lid", "cover", "top lid"],
            "Handle_Issues": ["handle", "grip", "holder"],
            "Gasket_Issues": ["gasket", "seal", "rubber seal", "ring"],
            "Whistle_Issues": ["whistle", "whistle sound", "whistles"],
            "Valve_Issues": ["valve", "safety valve", "pressure valve"],
            "Coating_Issues": ["coating", "non stick", "teflon", "non-stick", "nonstick", "peeling", "scratch"],
            "Build_Quality": ["sturdy", "solid", "flimsy", "durable", "build", "material", "body", "steel", "aluminium", "quality"],
            "Material_Quality": ["material", "plastic", "steel", "iron", "copper", "brass", "rubber"],
            "Overall_Quality": ["quality", "excellent", "superb", "pathetic", "useless", "terrible", "amazing"],
        },
        "Performance": {
            "Stopped_Working": ["stopped working", "not working", "doesn't work", "stopped", "dead", "defective"],
            "Cooking_Performance": ["cooking", "cook", "cooks", "cooked", "rice", "dal", "curry", "food"],
            "Heating_Performance": ["heat", "heating", "hot", "warm", "temperature", "flame", "burner", "fire"],
            "Induction_Performance": ["induction", "induction base", "compatible", "induction cooktop", "magnetic"],
            "Efficiency": ["fast", "quick", "slow", "efficient", "time", "saves time", "power"],
            "Flame_Gas": ["gas", "flame", "burner", "stove", "ignition", "lpg"],
        },
        "Usability": {
            "Ease_of_Use": ["easy", "simple", "convenient", "difficult", "complicated", "user friendly"],
            "Weight_Size": ["heavy", "light", "weight", "size", "compact", "large", "small", "big", "portable"],
            "Cleaning": ["clean", "cleaning", "wash", "easy clean", "dishwasher", "stain"],
        },
        "Value": {
            "Worth_Money": ["price", "value", "worth", "money", "expensive", "cheap", "cost", "affordable", "overpriced"],
            "Cheap_Quality": ["cheap quality", "cheap plastic", "low quality", "poor quality"],
            "Budget_Friendly": ["budget", "economical", "affordable", "best buy"],
        },
        "Delivery": {
            "Fast_Delivery": ["fast delivery", "quick delivery", "on time", "early", "prompt"],
            "Late_Delivery": ["late", "delayed", "delay", "waiting", "late delivery"],
            "Packaging_Quality": ["packaging", "packed", "box", "bubble wrap", "damaged delivery"],
            "Wrong_Product": ["wrong", "wrong product", "different product"],
            "Missing_Parts": ["missing", "parts missing", "incomplete", "accessories"],
        },
        "Customer Service": {
            "Poor_Service": ["service", "customer service", "support", "response", "helpline"],
            "Warranty_Issues": ["warranty", "replacement", "refund", "return", "exchange", "service center"],
            "Good_Service": ["helpful", "resolved", "replaced", "responsive"],
        },
        "Safety": {
            "Electrical_Safety": ["electric shock", "spark", "short circuit", "fire", "current"],
            "Physical_Safety": ["burn", "cut", "safe", "safety", "dangerous", "harm", "injury", "accident"],
        },
        "Features": {
            "Auto_Ignition": ["auto ignition", "ignition", "auto", "piezo"],
            "Induction_Compatible": ["induction compatible", "induction base", "induction friendly"],
            "Timer_Controls": ["timer", "control", "temperature control", "preset", "settings"],
            "Design_Look": ["design", "look", "beautiful", "stylish", "modern", "colour", "color", "shape"],
        },
        "Brand": {
            "Trust_Reputation": ["brand", "trust", "trusted", "reliable", "reputation", "original", "genuine", "authentic"],
            "Satisfaction": ["satisfied", "happy", "pleased", "love", "disappointed", "regret", "recommend"],
            "General_Feedback": ["okay", "ok", "fine", "average", "normal", "expected", "decent", "good", "nice", "great", "bad", "worst", "best"],
        },
    }
    
    # Auto-map discovered n-grams to categories
    enrichment = defaultdict(lambda: defaultdict(list))
    unmapped = []
    
    for ngram_type in ["bigrams", "trigrams"]:
        if ngram_type not in ngram_results:
            continue
        
        all_ngrams = set()
        for pair in ngram_results[ngram_type]["top_global"]:
            all_ngrams.add(pair[0])
        for pair in ngram_results[ngram_type]["top_negative"]:
            all_ngrams.add(pair[0])
        for pair in ngram_results[ngram_type]["top_positive"]:
            all_ngrams.add(pair[0])
        
        for ngram in all_ngrams:
            mapped = False
            best_cat = None
            best_subcat = None
            best_score = 0
            
            for cat, subcats in CATEGORY_PATTERNS.items():
                for subcat, keywords in subcats.items():
                    # Score based on keyword overlap
                    ngram_words = set(ngram.split())
                    for kw in keywords:
                        kw_words = set(kw.lower().split())
                        overlap = len(ngram_words & kw_words)
                        if overlap > 0:
                            score = overlap / max(len(ngram_words), len(kw_words))
                            if score > best_score:
                                best_score = score
                                best_cat = cat
                                best_subcat = subcat
                                mapped = True
            
            if mapped and best_score >= 0.3:
                enrichment[best_cat][best_subcat].append(ngram)
            elif not mapped:
                unmapped.append(ngram)
    
    # Print results
    total_mapped = sum(
        len(ngrams) for subcats in enrichment.values() for ngrams in subcats.values()
    )
    print(f"   Mapped: {total_mapped} n-grams")
    print(f"   Unmapped: {len(unmapped)} n-grams")
    
    for cat, subcats in sorted(enrichment.items()):
        print(f"\n   📂 {cat}:")
        for subcat, ngrams in sorted(subcats.items()):
            print(f"      {subcat} (+{len(ngrams)}): {ngrams[:5]}...")
    
    return dict(enrichment), unmapped


def merge_with_existing_taxonomy(enrichment: dict) -> dict:
    """Merge new keywords with existing FULL_TAXONOMY from category_classifier.py."""
    print("\n🔄 Merging with existing taxonomy...")
    
    # Load existing learned keywords if available
    learned_path = PROJECT_ROOT / "ml_pipeline" / "learned_keywords.json"
    existing_learned = {}
    if learned_path.exists():
        with open(learned_path, "r", encoding="utf-8") as f:
            existing_learned = json.load(f)
        print(f"   Loaded existing learned_keywords.json ({len(existing_learned)} categories)")
    
    # Merge
    merged = defaultdict(lambda: defaultdict(list))
    
    # Start with existing
    for cat, subcats in existing_learned.items():
        if isinstance(subcats, dict):
            for subcat, keywords in subcats.items():
                if isinstance(keywords, list):
                    merged[cat][subcat] = list(keywords)
    
    # Add new enrichment
    new_count = 0
    for cat, subcats in enrichment.items():
        for subcat, ngrams in subcats.items():
            existing = set(merged[cat][subcat])
            for ngram in ngrams:
                if ngram not in existing:
                    merged[cat][subcat].append(ngram)
                    new_count += 1
    
    print(f"   Added {new_count} new keywords across taxonomy")
    
    # Save enriched keywords
    output = {cat: dict(subcats) for cat, subcats in merged.items()}
    with open(TAXONOMY_OUTPUT, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"   💾 Saved enriched taxonomy to: {TAXONOMY_OUTPUT}")
    
    return output


def generate_good_bad_indicators(ngram_results: dict) -> tuple:
    """Extract good/bad sentiment indicators from positive/negative n-grams."""
    print("\n😊😠 Extracting sentiment indicators...")
    
    good_indicators = set()
    bad_indicators = set()
    
    for ngram_type in ["unigrams", "bigrams"]:
        if ngram_type not in ngram_results:
            continue
        
        for ngram, count in ngram_results[ngram_type]["top_positive"]:
            if count >= 15:
                good_indicators.add(ngram)
        
        for ngram, count in ngram_results[ngram_type]["top_negative"]:
            if count >= 15:
                bad_indicators.add(ngram)
    
    print(f"   Good indicators: {len(good_indicators)}")
    print(f"   Bad indicators: {len(bad_indicators)}")
    print(f"   Sample good: {sorted(good_indicators)[:10]}")
    print(f"   Sample bad: {sorted(bad_indicators)[:10]}")
    
    return sorted(good_indicators), sorted(bad_indicators)


def main():
    print("=" * 70)
    print("PHASE 2: ML TAXONOMY ENRICHMENT FROM REAL COMPETITOR DATA")
    print("=" * 70)
    
    # Step 1: Read all reviews from Excel
    reviews = read_excel_reviews()
    
    # Step 2: Run n-gram analysis
    ngram_results = analyze_ngrams(reviews)
    
    # Step 3: Map to taxonomy categories
    enrichment, unmapped = map_ngrams_to_taxonomy(ngram_results)
    
    # Step 4: Merge with existing taxonomy
    merged_taxonomy = merge_with_existing_taxonomy(enrichment)
    
    # Step 5: Extract sentiment indicators
    good_indicators, bad_indicators = generate_good_bad_indicators(ngram_results)
    
    # Step 6: Save comprehensive report
    report = {
        "stats": {
            "total_reviews_analyzed": len(reviews),
            "unique_bigrams": ngram_results.get("bigrams", {}).get("total_unique", 0),
            "unique_trigrams": ngram_results.get("trigrams", {}).get("total_unique", 0),
            "new_keywords_added": sum(
                len(ngrams) for subcats in enrichment.values() for ngrams in subcats.values()
            ),
            "unmapped_ngrams": len(unmapped),
        },
        "enrichment_by_category": enrichment,
        "unmapped_top_ngrams": unmapped[:50],
        "good_indicators": good_indicators,
        "bad_indicators": bad_indicators,
        "top_bigrams_global": ngram_results.get("bigrams", {}).get("top_global", [])[:50],
        "top_trigrams_global": ngram_results.get("trigrams", {}).get("top_global", [])[:50],
    }
    
    with open(REPORT_OUTPUT, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    
    print(f"\n{'=' * 70}")
    print("ENRICHMENT COMPLETE")
    print(f"{'=' * 70}")
    print(f"📊 Reviews analyzed: {len(reviews):,}")
    print(f"📖 New keywords discovered: {report['stats']['new_keywords_added']}")
    print(f"😊 Good indicators: {len(good_indicators)}")
    print(f"😠 Bad indicators: {len(bad_indicators)}")
    print(f"💾 Enriched taxonomy: {TAXONOMY_OUTPUT}")
    print(f"📄 Full report: {REPORT_OUTPUT}")


if __name__ == "__main__":
    main()
