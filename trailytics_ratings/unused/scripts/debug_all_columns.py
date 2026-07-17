"""Inspect EVERY column thoroughly, especially columns 0-3."""
import openpyxl

wb = openpyxl.load_workbook('dist/assets/prestige_comp_reviews_data.xlsx', read_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*80}")
    
    # Get header row
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)
        break
    
    print(f"ALL HEADERS ({len(headers)}):")
    for i, h in enumerate(headers):
        print(f"  Col[{i:2d}] = '{h}'")
    
    # Show 10 complete sample rows, different PIDs
    print(f"\n10 SAMPLE ROWS (all columns):")
    seen_pids = set()
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = str(row[4]) if row[4] else ""
        if pid in seen_pids:
            continue
        seen_pids.add(pid)
        count += 1
        if count > 10:
            break
        
        print(f"\n  --- PID: {pid} ---")
        for i, val in enumerate(row):
            header = headers[i] if i < len(headers) else f"Col{i}"
            val_str = str(val)[:120] if val else "(empty)"
            if val_str == "\\N":
                val_str = "(null)"
            print(f"    [{i:2d}] {header:25s} = {val_str}")

wb.close()
