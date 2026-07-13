"""Check ALL columns in Excel for brand/product data."""
import openpyxl

wb = openpyxl.load_workbook('dist/assets/prestige_comp_reviews_data.xlsx', read_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    
    # Get header row
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)
        break
    
    print(f"Headers ({len(headers)}):")
    for i, h in enumerate(headers):
        print(f"  Col[{i}] = {h}")
    
    # Show 3 full rows with ALL columns
    print(f"\nSample rows:")
    count = 0
    for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
        count += 1
        print(f"\n  --- Row {count} ---")
        for i, val in enumerate(row):
            header = headers[i] if i < len(headers) else f"Col{i}"
            val_str = str(val)[:100] if val else "(empty)"
            print(f"  [{i}] {header}: {val_str}")

wb.close()
