"""Inspect the competitor Excel file to understand data shape."""
import openpyxl
import json
from collections import Counter

wb = openpyxl.load_workbook(
    'C:/Users/monst/OneDrive/Documents/GitHub/Ratings/dist/assets/prestige_comp_reviews_data.xlsx',
    read_only=True
)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    
    samples = []
    null_text_count = 0
    short_text_count = 0
    total = 0
    pids = set()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        pid = str(row[4]) if row[4] and str(row[4]) != '\\N' else ''
        if pid:
            pids.add(pid)
        
        title = str(row[7]) if row[7] and str(row[7]) != '\\N' else ''
        body = str(row[8]) if row[8] and str(row[8]) != '\\N' else ''
        text = f"{title}. {body}".strip('. ') if title or body else ''
        
        if not text:
            null_text_count += 1
            continue
        if len(text) < 10:
            short_text_count += 1
        
        rating = int(row[6]) if row[6] and str(row[6]) != '\\N' else 0
        url = str(row[5])[:120] if row[5] else ''
        
        if len(samples) < 5 and len(text) > 50:
            samples.append({
                'pid': pid,
                'url': url,
                'rating': rating,
                'title': title[:80],
                'body': body[:150],
            })
    
    print(f"Total rows: {total}")
    print(f"Unique PIDs: {len(pids)}")
    print(f"Null text: {null_text_count}")
    print(f"Short text (<10 chars): {short_text_count}")
    print(f"\nSample reviews:")
    for i, s in enumerate(samples):
        print(f"\n  [{i+1}] PID={s['pid']} | Rating={s['rating']}")
        print(f"      URL: {s['url']}")
        print(f"      Title: {s['title']}")
        print(f"      Body: {s['body']}")

wb.close()
