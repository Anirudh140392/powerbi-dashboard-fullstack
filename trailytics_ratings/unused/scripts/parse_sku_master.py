"""
Parse SKU Master Excel → sku_master.json
Reads the TTK Prestige ECOM SKU master and outputs structured JSON
for dashboard consumption.
"""
import openpyxl
import json
import os
import sys

EXCEL_PATH = r"C:\Users\monst\Downloads\ECOM SKU_Platform Details 05032026_TTK Prestige.xlsx"
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "src", "data", "sku_master.json")

# Column mapping (1-indexed) based on Row 2 headers
COL_MAP = {
    "sl_no": 1,
    "business_segment": 2,
    "category": 3,
    "subcategory_l1": 4,
    "subcategory_l2": 5,
    "sku_code": 6,
    "model": 7,
    "asin": 8,        # Amazon ASIN
    "fsn": 9,         # Flipkart FSN
    "blinkit_ref": 10,
    "blinkit_pid": 11,
    "swiggy_ref": 12,
    "swiggy_pid": 13,
    "zepto_ref": 14,
    "bigbasket_ref": 15,
    "tatacliq_id": 16,
    "myntra_id": 17,
    "jiomart_id": 18,
    "ajio_id": 19,
    "status": 20,      # Pareto / Non-Pareto / NPD
    "mrp": 21,
    "mop": 22,
}

# Material extraction from Subcategory L1 and L2
MATERIAL_KEYWORDS = {
    "Aluminium": ["aluminium", "aluminum"],
    "Stainless Steel": ["stainless steel", "ss "],
    "Triply": ["triply", "tri-ply", "tri ply"],
    "Hard Anodised": ["hard anodised", "hard anodized", "ha "],
    "Cast Iron": ["cast iron"],
    "Non-Stick": ["non-stick", "nonstick", "non stick"],
    "Ceramic": ["ceramic"],
    "Glass": ["glass", "borosilicate"],
    "Copper": ["copper"],
}


def extract_material(subcat_l1: str, subcat_l2: str, category: str) -> str:
    """Extract material type from subcategory fields."""
    combined = f"{subcat_l1 or ''} {subcat_l2 or ''}".lower()
    for material, keywords in MATERIAL_KEYWORDS.items():
        for kw in keywords:
            if kw in combined:
                return material
    return "Other"


def extract_wattage(model: str, subcat_l1: str, subcat_l2: str) -> str | None:
    """Extract wattage for electrical appliances."""
    import re
    combined = f"{model or ''} {subcat_l1 or ''} {subcat_l2 or ''}"
    match = re.search(r'(\d{3,4})\s*[wW](?:att)?', combined)
    if match:
        return f"{match.group(1)}W"
    return None


def parse_excel():
    """Parse the Excel file and return structured data."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    skus = []
    for row in range(3, ws.max_row + 1):  # Data starts at row 3
        def cell(col_key):
            val = ws.cell(row, COL_MAP[col_key]).value
            return str(val).strip() if val is not None else None

        category = cell("category")
        if not category:
            continue

        subcat_l1 = cell("subcategory_l1")
        subcat_l2 = cell("subcategory_l2")
        model = cell("model")
        status = cell("status")

        # Normalize status
        if status:
            status_lower = status.lower().strip()
            if "pareto" in status_lower and "non" not in status_lower:
                status = "Pareto"
            elif "non" in status_lower:
                status = "Non-Pareto"
            elif "npd" in status_lower:
                status = "NPD"

        material = extract_material(subcat_l1, subcat_l2, category)

        # For electrical appliances, extract wattage
        wattage = None
        elec_cats = ["mixer grinder", "induction", "air fryer", "kettle", "iron",
                     "toaster", "otg", "wet grinder", "sandwich"]
        if category and any(ec in category.lower() for ec in elec_cats):
            wattage = extract_wattage(model or "", subcat_l1 or "", subcat_l2 or "")

        sku = {
            "skuCode": cell("sku_code"),
            "model": model,
            "category": category,
            "subcategoryL1": subcat_l1,
            "subcategoryL2": subcat_l2,
            "businessSegment": cell("business_segment"),
            "status": status,  # Pareto / Non-Pareto / NPD
            "material": material,
            "wattage": wattage,
            "mrp": float(cell("mrp")) if cell("mrp") else None,
            "mop": float(cell("mop")) if cell("mop") else None,
            "platformIds": {
                "amazon": cell("asin"),
                "flipkart": cell("fsn"),
                "blinkit": cell("blinkit_ref"),
                "swiggy": cell("swiggy_ref"),
                "zepto": cell("zepto_ref"),
                "bigbasket": cell("bigbasket_ref"),
                "tatacliq": cell("tatacliq_id"),
                "myntra": cell("myntra_id"),
                "jiomart": cell("jiomart_id"),
                "ajio": cell("ajio_id"),
            },
        }
        skus.append(sku)

    wb.close()
    return skus


def main():
    print(f"Parsing Excel: {EXCEL_PATH}")
    skus = parse_excel()
    print(f"Parsed {len(skus)} SKUs")

    # Stats
    from collections import Counter
    status_dist = Counter(s["status"] for s in skus if s["status"])
    cat_dist = Counter(s["category"] for s in skus)
    mat_dist = Counter(s["material"] for s in skus)
    print(f"Status: {dict(status_dist)}")
    print(f"Categories: {dict(cat_dist)}")
    print(f"Materials: {dict(mat_dist)}")

    # Count SKUs with Amazon ASIN
    asin_count = sum(1 for s in skus if s["platformIds"].get("amazon"))
    print(f"SKUs with Amazon ASIN: {asin_count}")

    # Write JSON
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(skus, f, indent=2, ensure_ascii=False)
    print(f"Written to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
