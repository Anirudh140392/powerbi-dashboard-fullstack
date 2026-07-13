"""
Competitor Data Processing Pipeline
=====================================
Parses competitor review data from Excel, classifies via ML,
generates competitor product catalog, and builds SKU mappings.

Usage:
    python scripts/process_competitor_excel.py
    python scripts/process_competitor_excel.py --input path/to/new_data.xlsx
"""

import os
import sys
import re
import json
import math
import hashlib
import argparse
from collections import Counter, defaultdict
from pathlib import Path
from datetime import datetime

# Add project root and ML pipeline
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "ml_pipeline"))

import openpyxl

# ============================================================
# Import ML pipeline modules
# ============================================================
from sentiment_analyzer import analyze_sentiment_local, get_polarity_score
from category_classifier import (
    classify_all_labels, classify_with_rules_v7, get_sentiment,
    detect_indirect_sentiment, FULL_TAXONOMY
)

# ============================================================
# BRAND DETECTION CONFIG — from competitor_products.json + expanded
# ============================================================
# Priority-ordered: longer names first to avoid partial matches
KNOWN_BRANDS = [
    # Multi-word brands (check first)
    "morphy richards", "morphy-richards",
    "black decker", "black+decker", "black & decker",
    "euro guard", "euro-guard",
    "glen india", "glen appliances",
    "kent ro", "kent atta",
    "stovekraft", "stove kraft",
    "ttk prestige",
    # Single-word brands
    "prestige", "hawkins", "pigeon", "butterfly", "preethi", "bajaj",
    "philips", "wonderchef", "bosch", "panasonic", "sujata", "maharaja",
    "usha", "crompton", "havells", "orient", "kenstar", "inalsa", "glen",
    "vinod", "nirlep", "cello", "elica", "faber", "kaff", "hindware",
    "sunflame", "milton", "signora", "ganesh", "sumeet", "vidiem",
    "premier", "futura", "prestige", "hamilton", "cuisinart", "kitchenaid",
    "thermador", "whirlpool", "lg", "samsung", "ifb", "godrej", "voltas",
    "kent", "aquaguard", "livpure", "eureka forbes",
]

# Product category keywords for detection from review text or URL slug
PRODUCT_CATEGORY_PATTERNS = {
    "Pressure Cooker": [
        "pressure cooker", "prestige cooker", "outer lid", "inner lid",
        "cooker", "whistle", "5 litre", "3 litre", "2 litre",
        "aluminium cooker", "hard anodised cooker", "stainless steel cooker",
        "contura", "svachh", "popular", "deluxe", "nakshatra", "clip on",
    ],
    "Gas Stove": [
        "gas stove", "gas burner", "glass top", "stainless steel stove",
        "2 burner", "3 burner", "4 burner", "cooktop", "hob",
        "auto ignition", "brass burner", "drip tray",
    ],
    "Induction Cooktop": [
        "induction cooktop", "induction stove", "induction plate",
        "induction heater", "induction cook", "pic", "induction",
    ],
    "Mixer Grinder": [
        "mixer grinder", "mixer", "grinder", "juicer mixer",
        "wet grinder", "dry grinder", "chutney jar", "liquidizer",
        "blender", "750w", "550w", "500w", "1000w",
    ],
    "Non-Stick Cookware": [
        "non stick", "nonstick", "non-stick", "tawa", "kadhai", "kadai",
        "fry pan", "frying pan", "dosa tawa", "roti tawa", "omni tawa",
        "saucepan", "sauce pan", "milk pan",
    ],
    "Rice Cooker": [
        "rice cooker", "electric rice", "delight rice",
    ],
    "Electric Kettle": [
        "electric kettle", "kettle", "hot water",
    ],
    "Air Fryer": [
        "air fryer", "airfryer", "oil free",
    ],
    "Sandwich Maker": [
        "sandwich maker", "sandwich toaster", "grill sandwich",
    ],
    "OTG / Oven": [
        "otg", "oven toaster", "microwave", "oven",
    ],
    "Kitchen Tools": [
        "kitchen tool", "ladle", "spoon", "turner", "whisk",
        "peeler", "tong", "spatula", "serving spoon",
    ],
    "Stainless Steel Cookware": [
        "stainless steel", "triply", "tri-ply", "tri ply",
        "steel pan", "steel kadhai", "steel pot",
    ],
}


def detect_brand_from_text(text: str, url: str = "") -> str:
    """Detect brand from URL slug and/or review text."""
    combined = f"{url} {text}".lower()
    
    # Extract URL slug (the part before /dp/ on Amazon)
    url_slug = ""
    if "/dp/" in combined:
        parts = combined.split("/dp/")
        url_slug = parts[0].split("/")[-1] if parts else ""
        url_slug = url_slug.replace("-", " ").replace("_", " ")
    
    search_text = f"{url_slug} {combined}"
    
    for brand in KNOWN_BRANDS:
        # Search for brand with word boundaries
        pattern = r'\b' + re.escape(brand) + r'\b'
        if re.search(pattern, search_text):
            # Normalize brand name (capitalize)
            normalized = {
                "morphy richards": "Morphy Richards",
                "morphy-richards": "Morphy Richards",
                "black decker": "Black & Decker",
                "black+decker": "Black & Decker",
                "black & decker": "Black & Decker",
                "ttk prestige": "Prestige",
                "stovekraft": "Pigeon",
                "stove kraft": "Pigeon",
                "eureka forbes": "Eureka Forbes",
                "glen india": "Glen",
                "glen appliances": "Glen",
                "kent ro": "Kent",
                "kent atta": "Kent",
            }
            return normalized.get(brand, brand.title())
    
    return "Unknown"


def detect_category_from_text(text: str, url: str = "") -> str:
    """Detect product category from review text and URL."""
    combined = f"{url} {text}".lower()
    
    # Score each category
    scores = {}
    for category, keywords in PRODUCT_CATEGORY_PATTERNS.items():
        score = 0
        for kw in keywords:
            if kw in combined:
                # Longer keywords get higher score
                score += len(kw.split())
        if score > 0:
            scores[category] = score
    
    if scores:
        return max(scores, key=scores.get)
    return "General"


def extract_product_name_from_url(url: str) -> str:
    """Try to extract a human-readable product name from URL slug."""
    if not url:
        return ""
    
    # Amazon: extract slug before /dp/
    if "/dp/" in url:
        parts = url.split("/dp/")
        slug = parts[0].split("/")[-1] if parts else ""
        # Clean up the slug
        name = slug.replace("-", " ").replace("_", " ")
        # Remove generic words
        name = re.sub(r'\b(amazon|flipkart|www|com|in|http|https)\b', '', name, flags=re.IGNORECASE)
        name = re.sub(r'\s+', ' ', name).strip()
        if len(name) > 5:
            return name.title()
    
    # Flipkart: extract from path segments
    if "flipkart" in url.lower():
        path = url.split("?")[0]
        segments = [s for s in path.split("/") if s and s not in ["abc", "p", "xyz", ""]]
        if segments:
            name = segments[-1].replace("-", " ")
            if len(name) > 5:
                return name.title()
    
    return ""


def generate_review_id(platform: str, pid: str, idx: int) -> str:
    """Generate a unique review ID."""
    prefix = "AMZ" if platform == "amazon" else "FK"
    pid_short = pid[:12] if pid else "UNK"
    return f"{prefix}-{pid_short}-{idx:04d}"


def read_and_parse_excel(excel_path: str) -> list:
    """Phase 1: Read Excel and parse all reviews into structured format."""
    print(f"\n{'='*60}")
    print("PHASE 1: PARSING EXCEL FILE")
    print(f"{'='*60}")
    print(f"📂 Source: {excel_path}")
    
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    
    all_reviews = []
    product_registry = defaultdict(lambda: {
        "count": 0, "ratings": [], "brands": Counter(), "categories": Counter(),
        "urls": set(), "platform": "",
    })
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        platform = sheet_name.lower()
        pid_counter = Counter()
        skipped = 0
        
        print(f"\n   📋 Sheet: {sheet_name}")
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            pid = str(row[4]) if row[4] and str(row[4]) != "\\N" else ""
            if not pid:
                skipped += 1
                continue
            
            title = str(row[7]) if row[7] and str(row[7]) != "\\N" else ""
            body = str(row[8]) if row[8] and str(row[8]) != "\\N" else ""
            text = f"{title}. {body}".strip(". ")
            
            if not text or len(text) < 3:
                skipped += 1
                continue
            
            # Parse rating
            rating = 0
            try:
                rating = int(row[6]) if row[6] and str(row[6]) != "\\N" else 0
            except (ValueError, TypeError):
                pass
            
            # Parse date
            date_str = ""
            if row[10]:
                try:
                    if isinstance(row[10], datetime):
                        date_str = row[10].strftime("%Y-%m-%d")
                    else:
                        date_str = str(row[10])[:10]
                except Exception:
                    date_str = ""
            
            url = str(row[5])[:500] if row[5] and str(row[5]) != "\\N" else ""
            reviewer = str(row[12]) if row[12] and str(row[12]) != "\\N" else ""
            helpful = 0
            try:
                helpful = int(row[16]) if row[16] and str(row[16]) != "\\N" else 0
            except (ValueError, TypeError):
                pass
            
            review_type = str(row[15]) if row[15] and str(row[15]) != "\\N" else ""
            verified = review_type.lower() == "review"
            
            # Detect brand and category
            brand = detect_brand_from_text(text, url)
            category = detect_category_from_text(text, url)
            product_name = extract_product_name_from_url(url)
            
            # Track per-PID statistics
            pid_counter[pid] += 1
            idx = pid_counter[pid]
            
            # Build product registry
            reg = product_registry[pid]
            reg["count"] += 1
            reg["ratings"].append(rating)
            reg["brands"][brand] += 1
            reg["categories"][category] += 1
            reg["urls"].add(url[:200])
            reg["platform"] = platform
            
            review = {
                "reviewId": generate_review_id(platform, pid, idx),
                "productId": pid,
                "productName": product_name if product_name else f"{brand} {category}".strip(),
                "brand": brand,
                "category": category,
                "rating": rating,
                "text": text,
                "date": date_str,
                "verified": verified,
                "platform": platform,
                "reviewer": reviewer,
                "helpfulCount": helpful,
            }
            
            all_reviews.append(review)
        
        print(f"      Reviews loaded: {len([r for r in all_reviews if r['platform'] == platform]):,}")
        print(f"      Skipped (empty): {skipped}")
    
    wb.close()
    
    # Finalize product names using majority voting
    print(f"\n   📊 Finalizing product identities via majority voting...")
    pid_brand_map = {}
    pid_category_map = {}
    pid_name_map = {}
    
    for pid, reg in product_registry.items():
        # Use most common brand for this PID
        best_brand = reg["brands"].most_common(1)[0][0] if reg["brands"] else "Unknown"
        best_category = reg["categories"].most_common(1)[0][0] if reg["categories"] else "General"
        
        pid_brand_map[pid] = best_brand
        pid_category_map[pid] = best_category
        
        # Try to get a good product name from URLs
        for url in reg["urls"]:
            name = extract_product_name_from_url(url)
            if name and len(name) > 5:
                pid_name_map[pid] = name
                break
        if pid not in pid_name_map:
            pid_name_map[pid] = f"{best_brand} {best_category}".strip()
    
    # Apply majority-voted brand/category/name to all reviews of same PID
    for review in all_reviews:
        pid = review["productId"]
        review["brand"] = pid_brand_map.get(pid, review["brand"])
        review["category"] = pid_category_map.get(pid, review["category"])
        if not review["productName"] or review["productName"] == f"{review['brand']} {review['category']}":
            review["productName"] = pid_name_map.get(pid, review["productName"])
    
    print(f"   Total reviews: {len(all_reviews):,}")
    print(f"   Unique products (PIDs): {len(product_registry)}")
    
    # Brand distribution
    brand_dist = Counter(r["brand"] for r in all_reviews)
    print(f"\n   📊 Brand Distribution:")
    for brand, count in brand_dist.most_common(20):
        pct = count / len(all_reviews) * 100
        print(f"      {brand:20s} {count:6,} ({pct:5.1f}%)")
    
    # Category distribution
    cat_dist = Counter(r["category"] for r in all_reviews)
    print(f"\n   📂 Category Distribution:")
    for cat, count in cat_dist.most_common():
        pct = count / len(all_reviews) * 100
        print(f"      {cat:25s} {count:6,} ({pct:5.1f}%)")
    
    return all_reviews, product_registry


def run_ml_classification(reviews: list) -> list:
    """Phase 3: Run ML pipeline on all reviews using v8 multi-label classifier."""
    print(f"\n{'='*60}")
    print("PHASE 3: ML CLASSIFICATION (v8 Multi-Label)")
    print(f"{'='*60}")
    print(f"   Processing {len(reviews):,} reviews through ML pipeline...")
    
    # Load enriched keywords if available
    enriched_path = PROJECT_ROOT / "ml_pipeline" / "enriched_keywords.json"
    if enriched_path.exists():
        with open(enriched_path, "r", encoding="utf-8") as f:
            enriched = json.load(f)
        print(f"   ✅ Loaded enriched taxonomy ({sum(len(s) for c in enriched.values() for s in c.values())} extra keywords)")
    
    classified = []
    stats = {
        "sentiments": Counter(),
        "categories": Counter(),
        "subcategories": Counter(),
        "low_confidence": 0,
        "multi_label_count": Counter(),  # How many labels per review
        "indirect_sentiment": Counter(),
        "negated_count": 0,
        "high_impact_issues": Counter(),
    }
    
    # Process in batches with progress
    batch_size = 5000
    total = len(reviews)
    
    for i in range(0, total, batch_size):
        batch = reviews[i:i+batch_size]
        batch_num = i // batch_size + 1
        total_batches = math.ceil(total / batch_size)
        print(f"   Batch {batch_num}/{total_batches} ({i:,}-{min(i+batch_size, total):,})...")
        
        for review in batch:
            text = review["text"]
            rating = review["rating"]
            product = review.get("productName", "")
            
            # 1. Sentiment (local — TextBlob + VADER)
            sentiment, sent_confidence, is_complex = analyze_sentiment_local(text)
            polarity = get_polarity_score(text)
            
            # 2. Indirect sentiment detection
            indirect = detect_indirect_sentiment(text)
            if indirect:
                stats["indirect_sentiment"][indirect] += 1
            
            # 3. Override sentiment if rating strongly disagrees
            if rating >= 4 and sentiment == "NEGATIVE" and sent_confidence < 0.7:
                sentiment = "POSITIVE"
                sent_confidence = 0.6
            elif rating <= 2 and sentiment == "POSITIVE" and sent_confidence < 0.7:
                sentiment = "NEGATIVE"
                sent_confidence = 0.6
            elif rating == 3 and indirect:
                # For 3-star reviews, use indirect sentiment if detected
                sentiment = indirect
                sent_confidence = 0.55
            
            # 4. v8 Multi-label classification
            all_labels = classify_all_labels(text, rating, product)
            
            # Primary label (highest impact * confidence)
            primary = all_labels[0] if all_labels else {
                "category": "General", "subcategory": "General_Feedback",
                "confidence": 0.3, "impact": 0.2, "keywords_matched": [],
                "is_negated": False, "sentiment_direction": "neutral"
            }
            
            # 5. Build enriched review with multi-label data
            review["sentiment"] = sentiment
            review["sentimentConfidence"] = round(sent_confidence, 3)
            review["polarity"] = round(polarity, 3)
            review["sentimentCategory"] = primary["category"]
            review["subcategory"] = primary["subcategory"]
            review["categoryConfidence"] = round(primary["confidence"], 3)
            
            # v8 additions: multi-label, impact, negation
            review["impact"] = primary["impact"]
            review["sentimentDirection"] = primary["sentiment_direction"]
            review["isNegated"] = primary["is_negated"]
            
            # Store ALL matched labels (for dashboard drill-down)
            review["allLabels"] = [
                {
                    "category": l["category"],
                    "subcategory": l["subcategory"],
                    "confidence": round(l["confidence"], 3),
                    "impact": l["impact"],
                    "sentimentDirection": l["sentiment_direction"],
                    "isNegated": l["is_negated"],
                    "keywords": l["keywords_matched"][:3],
                }
                for l in all_labels[:5]  # Top 5 labels per review
            ]
            
            if indirect:
                review["indirectSentiment"] = indirect
            
            classified.append(review)
            
            # Stats
            stats["sentiments"][sentiment] += 1
            stats["categories"][primary["category"]] += 1
            stats["subcategories"][primary["subcategory"]] += 1
            stats["multi_label_count"][min(len(all_labels), 5)] += 1
            if primary["confidence"] < 0.5:
                stats["low_confidence"] += 1
            if primary["is_negated"]:
                stats["negated_count"] += 1
            if primary["impact"] >= 0.7:
                stats["high_impact_issues"][primary["subcategory"]] += 1
    
    # Print comprehensive summary
    print(f"\n   ✅ Classification complete!")
    
    print(f"\n   😊 Sentiment Distribution:")
    for sent, count in stats["sentiments"].most_common():
        pct = count / total * 100
        print(f"      {sent:10s} {count:6,} ({pct:5.1f}%)")
    
    print(f"\n   📂 Primary Category Distribution:")
    for cat, count in stats["categories"].most_common():
        pct = count / total * 100
        bar = "█" * int(pct / 3)
        print(f"      {cat:25s} {count:6,} ({pct:5.1f}%) {bar}")
    
    print(f"\n   🏷️  Multi-label Distribution:")
    for n, count in sorted(stats["multi_label_count"].items()):
        pct = count / total * 100
        print(f"      {n} labels: {count:6,} ({pct:5.1f}%)")
    
    print(f"\n   🔍 Indirect Sentiment: {dict(stats['indirect_sentiment'])}")
    print(f"   ⚡ Negated reviews: {stats['negated_count']:,}")
    
    print(f"\n   🚨 High-Impact Issues (impact ≥ 0.7):")
    for subcat, count in stats["high_impact_issues"].most_common(15):
        print(f"      {subcat:25s} {count:6,}")
    
    print(f"\n   ⚠️  Low confidence (<0.5): {stats['low_confidence']:,} ({stats['low_confidence']/total*100:.1f}%)")
    
    return classified


def build_competitor_products(product_registry: dict, reviews: list) -> list:
    """Phase 4a: Build competitor_products.json from real data."""
    print(f"\n{'='*60}")
    print("PHASE 4: BUILDING PRODUCT CATALOG & SKU MAPPINGS")
    print(f"{'='*60}")
    
    # Aggregate per-PID data
    pid_reviews = defaultdict(list)
    for r in reviews:
        pid_reviews[r["productId"]].append(r)
    
    products = []
    for pid, reg in product_registry.items():
        if reg["count"] < 3:  # Skip PIDs with fewer than 3 reviews
            continue
        
        revs = pid_reviews.get(pid, [])
        best_brand = reg["brands"].most_common(1)[0][0] if reg["brands"] else "Unknown"
        best_category = reg["categories"].most_common(1)[0][0] if reg["categories"] else "General"
        avg_rating = round(sum(reg["ratings"]) / len(reg["ratings"]), 1) if reg["ratings"] else 0
        
        # Get best product name
        product_name = f"{best_brand} {best_category}"
        for url in reg["urls"]:
            name = extract_product_name_from_url(url)
            if name and len(name) > 8:
                product_name = name
                break
        
        # Generate stable product ID
        prefix = best_brand[:3].upper() if best_brand != "Unknown" else "UNK"
        pid_hash = hashlib.md5(pid.encode()).hexdigest()[:4].upper()
        product_id = f"{prefix}-{pid_hash}"
        
        # Get URL
        url = ""
        for u in reg["urls"]:
            if u:
                url = u
                break
        
        products.append({
            "productId": product_id,
            "name": product_name,
            "brand": best_brand,
            "category": best_category,
            "platformId": pid,
            "platform": reg["platform"],
            "url": url,
            "rating": avg_rating,
            "reviewCount": reg["count"],
        })
    
    # Sort by review count desc
    products.sort(key=lambda p: -p["reviewCount"])
    
    print(f"   Products catalog: {len(products)} products (min 3 reviews)")
    print(f"   Top brands by products:")
    brand_prods = Counter(p["brand"] for p in products)
    for brand, count in brand_prods.most_common(15):
        print(f"      {brand:20s} {count:4} products")
    
    return products


def build_sku_mappings(competitor_products: list) -> list:
    """Phase 4b: Build competitor ↔ own SKU mappings."""
    print(f"\n   🔗 Building competitor ↔ own SKU mappings...")
    
    # Load own product data
    own_reviews_path = PROJECT_ROOT / "src" / "data" / "processed_reviews.json"
    own_products = set()
    own_by_category = defaultdict(list)
    
    if own_reviews_path.exists():
        try:
            with open(own_reviews_path, "r", encoding="utf-8") as f:
                own_reviews = json.load(f)
            for r in own_reviews:
                product = r.get("product", "")
                if product:
                    own_products.add(product)
            
            # Categorize own products
            for product in own_products:
                p_lower = product.lower()
                for category, keywords in PRODUCT_CATEGORY_PATTERNS.items():
                    if any(kw in p_lower for kw in keywords):
                        own_by_category[category].append(product)
                        break
                else:
                    own_by_category["Other"].append(product)
            
            print(f"      Own products loaded: {len(own_products)}")
            for cat, prods in sorted(own_by_category.items()):
                print(f"         {cat}: {len(prods)} products")
        except Exception as e:
            print(f"      ⚠️  Could not load own reviews: {e}")
    
    # Build mappings: each competitor SKU → matching own SKUs by category
    mappings = []
    for comp in competitor_products:
        category = comp["category"]
        matching_own = own_by_category.get(category, [])
        
        if matching_own:
            mappings.append({
                "competitorPlatformId": comp["platformId"],
                "competitorProductId": comp["productId"],
                "competitorName": comp["name"],
                "competitorBrand": comp["brand"],
                "category": category,
                "platform": comp["platform"],
                "ownProducts": matching_own[:10],  # Top 10 own products in same category
            })
    
    print(f"      SKU mappings generated: {len(mappings)}")
    return mappings


def save_outputs(reviews: list, products: list, mappings: list):
    """Save all output files."""
    print(f"\n{'='*60}")
    print("SAVING OUTPUT FILES")
    print(f"{'='*60}")
    
    # 1. competitor_reviews.json
    output_path = PROJECT_ROOT / "src" / "data" / "competitor_reviews.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(reviews, f, ensure_ascii=False, indent=2)
    size_mb = output_path.stat().st_size / (1024 * 1024)
    print(f"   💾 {output_path.name}: {len(reviews):,} reviews ({size_mb:.1f} MB)")
    
    # 2. competitor_products.json
    products_path = PROJECT_ROOT / "src" / "data" / "competitor_products.json"
    with open(products_path, "w", encoding="utf-8") as f:
        json.dump({"competitors": products}, f, ensure_ascii=False, indent=2)
    print(f"   💾 {products_path.name}: {len(products)} products")
    
    # 3. competitor_sku_mappings.json
    mappings_path = PROJECT_ROOT / "src" / "data" / "competitor_sku_mappings.json"
    with open(mappings_path, "w", encoding="utf-8") as f:
        json.dump(mappings, f, ensure_ascii=False, indent=2)
    print(f"   💾 {mappings_path.name}: {len(mappings)} mappings")


def main():
    parser = argparse.ArgumentParser(description="Process competitor review data from Excel")
    parser.add_argument("--input", "-i",
                        default=str(PROJECT_ROOT / "dist" / "assets" / "prestige_comp_reviews_data.xlsx"),
                        help="Path to Excel file")
    parser.add_argument("--sample", "-s", type=int, default=None,
                        help="Process only N reviews for testing")
    args = parser.parse_args()
    
    print("=" * 60)
    print("COMPETITOR DATA PROCESSING PIPELINE")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    # Phase 1: Parse Excel
    reviews, product_registry = read_and_parse_excel(args.input)
    
    if args.sample:
        reviews = reviews[:args.sample]
        print(f"\n   ⚠️  Sampling {args.sample} reviews for testing")
    
    # Phase 3: ML Classification
    classified = run_ml_classification(reviews)
    
    # Phase 4: Build product catalog and SKU mappings
    products = build_competitor_products(product_registry, classified)
    mappings = build_sku_mappings(products)
    
    # Save all outputs
    save_outputs(classified, products, mappings)
    
    print(f"\n{'='*60}")
    print("PIPELINE COMPLETE")
    print(f"Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
